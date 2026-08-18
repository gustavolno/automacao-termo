import openpyxl
from datetime import datetime
from pathlib import Path
from gerador_boletos.config import PLANILHA_ENTRADA
from gerador_boletos.logger import log

class ExcelService:
    def __init__(self, caminho_planilha=PLANILHA_ENTRADA):
        self.caminho = caminho_planilha
        self.wb = None
        self.sheet = None
        self.colunas = {} # Mapeia o nome da coluna para o índice (1-based)
    
    def carregar(self):
        """Carrega a planilha na memória e mapeia as colunas."""
        if not Path(self.caminho).exists():
            log.error(f"Planilha não encontrada: {self.caminho}")
            raise FileNotFoundError(f"Planilha não encontrada: {self.caminho}")
            
        self.wb = openpyxl.load_workbook(self.caminho)
        self.sheet = self.wb.active
        
        # Mapeia cabeçalhos da primeira linha
        for idx, cell in enumerate(self.sheet[1], start=1):
            if cell.value:
                self.colunas[cell.value.strip().upper()] = idx
                
        log.info(f"Planilha '{self.caminho}' carregada. {len(self.colunas)} colunas mapeadas.")

    def salvar(self):
        """Salva as alterações feitas no arquivo Excel."""
        if self.wb:
            try:
                self.wb.save(self.caminho)
            except PermissionError:
                log.error(f"Sem permissão para salvar {self.caminho}. O arquivo pode estar aberto em outro programa.")
                raise

    def get_linha_como_dict(self, numero_linha: int) -> dict:
        """Lê os valores de uma linha específica e os retorna como um dicionário indexado pelos cabeçalhos."""
        linha_dict = {'_linha_excel': numero_linha}
        for nome_coluna, idx_coluna in self.colunas.items():
            linha_dict[nome_coluna] = self.sheet.cell(row=numero_linha, column=idx_coluna).value
        return linha_dict

    def atualizar_celula(self, numero_linha: int, nome_coluna: str, valor):
        """Atualiza o valor de uma célula específica baseada no nome da coluna."""
        nome_coluna = nome_coluna.strip().upper()
        if nome_coluna not in self.colunas:
            # Adiciona a nova coluna no cabeçalho
            nova_idx = len(self.colunas) + 1
            self.sheet.cell(row=1, column=nova_idx).value = nome_coluna
            self.colunas[nome_coluna] = nova_idx
        
        idx_coluna = self.colunas[nome_coluna]
        self.sheet.cell(row=numero_linha, column=idx_coluna).value = valor

    def listar_registros_elegiveis(self):
        """
        Gera uma lista de dicionários para os registros que estão PENDENTES
        ou que foram PROCESSANDO/GERADOS mas não tiveram o email ENVIADO.
        """
        if not self.sheet:
            self.carregar()
            
        registros = []
        # Começa da linha 2 (ignorando cabeçalho)
        for row_idx in range(2, self.sheet.max_row + 1):
            status_boleto = str(self.sheet.cell(row=row_idx, column=self.colunas.get('STATUS_BOLETO', 0)).value or '').strip().upper()
            status_email = str(self.sheet.cell(row=row_idx, column=self.colunas.get('STATUS_EMAIL', 0)).value or '').strip().upper()
            
            # Pula linhas completamente vazias (sem ID ou Nome)
            id_val = self.sheet.cell(row=row_idx, column=self.colunas.get('ID', 0)).value
            if not id_val:
                continue

            # Regras de elegibilidade descritas na documentação:
            # - Boleto pendente
            # - Boleto gerado mas email com erro/pendente
            if status_boleto == 'PENDENTE':
                registros.append(self.get_linha_como_dict(row_idx))
            elif status_boleto == 'GERADO' and status_email != 'ENVIADO':
                registros.append(self.get_linha_como_dict(row_idx))
                
        return registros
